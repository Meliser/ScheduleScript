#ввод данных о месяцах
monthsSize = int(input('Введите количество месяцев: '))
months = []
monthsTitles = []
for i in range(monthsSize):
    months.append(int(input('Количество дней: ')))
    monthsTitles.append(input('Название месяца: '))
print("\n")

from openpyxl import * 

#загрузка базы данных
print('Загрузка базы данных...',end='')
databaseBook = load_workbook('база_данных.xlsx')
databaseSheet = databaseBook['Лист1']
print('Завершено')

#получение счётчиков
cellsCounters = databaseSheet[databaseSheet['G7'].value : databaseSheet['H7'].value]

#создание списка из счетчиков
counters = []
for c in cellsCounters:
   counters.append(c[0].value)
   
#получение фамилий и комнат
surnamesRooms = databaseSheet[databaseSheet['G10'].value : databaseSheet['H10'].value]

#создание списка формата (фамилия,комната)
surnamesRoomsList = []
for c in surnamesRooms:
   surnamesRoomsList.append((c[0].value,c[1].value)) 
   
#создание списка формата [id,счетчик]
idCountersList = [ [id,counters[id]] for id in range(45)] #45 - количество людей на этаже

#загрузка графика 
sheduleBook = load_workbook('Графики_4_этаж.xlsx')
source = sheduleBook['Шаблон']

#создание графика на месяц
for i in range(monthsSize):
    #создание нового листа
    target = sheduleBook.copy_worksheet(source)
    target.title = monthsTitles[i]
    #запись названия месяца в ячейку 
    target.cell(column = 28,row = 2,value = monthsTitles[i] + ' 2019 г.')
    #сортировка списка по счетчикам
    idCountersList.sort(key=lambda i: i[1])
   
    print('Создание графика на ' + monthsTitles[i] + '...',end='')
    for row in range(months[i]):
        #запись фамилии
        target.cell(column=1, row=row+5, value=surnamesRoomsList[idCountersList[row][0]][0])
        #запись комнаты
        target.cell(column=2, row=row+5, value=surnamesRoomsList[idCountersList[row][0]][1])
        #запись дня
        target.cell(column = row+3,row = 4,value=row+1)
        #увеличение счетчика 
        idCountersList[row][1]+=1
    print('Завершено')
    
#сохранение графика
print('Сохранение графика...',end='')        
sheduleBook.save('Графики_4_этаж.xlsx')
print('Завершено')

#запись крайнего месяца, для которого составлен график
databaseSheet['H13'].value = monthsTitles[monthsSize-1]

#сортировка списка по id 
idCountersList.sort(key=lambda i: i[0])

#запись списка счетчиков обратно в базу данных
print('Запись счётчиков обратно в базу данных...',end='')
index = 0
for c in cellsCounters:
   c[0].value = idCountersList[index][1]
   index+=1
print('Завершено')  
 
#сохранение базы данных   
print('Сохранение базы данных...',end='') 
databaseBook.save('база_данных.xlsx')
print('Завершено')




